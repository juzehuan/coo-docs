"""Excel(.xlsx) 表格导出。

导出的表格由 COO、审计员与外部核查方直接用 Excel/WPS 打开。此前一律输出 CSV，
虽然能打开，但列宽、冻结表头、筛选都没有，长数字串还会被 Excel 自作主张地转成
科学计数法或抹掉前导零——订单号、MD5 这类值一旦被改写，导出的表格就与系统里的
记录对不上，而核查方并不知道是 Excel 干的。

**公式注入**：xlsx 同样有这个问题，而且比 CSV 更直接——openpyxl 见到以 `=` 开头
的字符串会把单元格类型判为公式(data_type='f')，Excel 打开即执行。这里所有单元格
一律强制为字符串类型(data_type='s')，既挡住 `=HYPERLINK`/`=cmd|...`，也顺带保住
长数字串的原样(前导零、精度都不丢)。因此**不需要**再做 csv_safe 那种加单引号的
中和——那个单引号在 Excel 里会显示出来，反而污染内容。

**write_only 模式**：普通模式下每个单元格都是一个 Python 对象，实测导出 10 万行
（审计导出的上限）峰值 RSS 达 302MB，而生成的文件只有 2.8MB——内存放大约 100 倍。
两个审计员同时导出就是 600MB，与第 54 轮 ZIP 在内存里整包构建属同一类问题。
改用 write_only 后同样 10 万行只需 49MB（**降低 91%**），文件与特性不变，代价是
慢约 3 秒（33s→36s）——用可控的时间换掉一个会把进程顶爆的内存尖峰，值得。

注意：write_only 模式下 `freeze_panes` 必须在写入任何行**之前**设置，写完再设
不会生效（实测回读为 None）。
"""
import io
import re
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 列宽上限：MD5、说明这类长文本不至于把整张表撑得没法看
MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 8

_HEADER_FILL = PatternFill("solid", fgColor="16263F")   # 与界面主色一致
_HEADER_FONT = Font(color="F5F1E4", bold=True)


# xlsx 是 XML：这些控制字符在 XML 1.0 里根本不合法，openpyxl 会直接抛
# IllegalCharacterError。制表符(\t)、换行(\n)、回车(\r)是合法的，不要动。
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _cell_text(v) -> str:
    """单元格取值，并剔除 XML 不允许的控制字符。

    **为什么必须在这里兜底**：附件原名来自用户上传,而 `safe_original_name`
    此前只截长度、不过滤控制字符——上传一个名字里带 NUL 的文件即可入库。
    该原名随后会进两处 xlsx：ZIP 里的交付清单,以及审计日志的 `target` 列。
    第 69 轮实测后果：**订单 ZIP 导出 500,审计导出也 500**。
    审计导出是全系统的合规记录,也就是说**一个用户上传的畸形文件名,
    会让所有人的审计导出永久失效**,而错误只是一个裸 500。

    替换为 U+FFFD 而不是直接删掉：交付清单列的是"附件原名",
    静默抹掉字节会让清单与实际文件名对不上；留一个替换符至少能看出
    这里原本有个不可打印字符。

    上传侧也已同步过滤（uploads.safe_original_name），但库里已有的记录
    只能靠这里兜住——这正是"中心化净化"必须存在的理由。
    """
    if v is None:
        return ""
    return _ILLEGAL_XLSX_RE.sub("\ufffd", str(v))


def build_xlsx(header: Sequence[str], rows: Iterable[Sequence],
               sheet_title: str = "Sheet1") -> bytes:
    """生成单表 xlsx 字节流（流式写入，内存与行数基本无关）。"""
    wb = Workbook(write_only=True)
    # 工作表名不得含 : \ / ? * [ ]，且不超过 31 字符
    ws = wb.create_sheet(_safe_title(sheet_title))
    # 必须在写入任何行之前设置，否则不生效
    ws.freeze_panes = "A2"

    widths = [len(_cell_text(h)) for h in header]
    head_cells = []
    for name in header:
        c = WriteOnlyCell(ws, value=_cell_text(name))
        c.data_type = "s"
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(vertical="center")
        head_cells.append(c)
    ws.append(head_cells)

    n = 0
    for n, row in enumerate(rows, start=1):
        cells = []
        for i, val in enumerate(row):
            text = _cell_text(val)
            c = WriteOnlyCell(ws, value=text)
            # 一律按字符串存：既阻断公式注入，也避免 Excel 改写长数字串
            c.data_type = "s"
            if i < len(widths):
                widths[i] = max(widths[i], len(text))
            else:
                widths.append(len(text))
            cells.append(c)
        ws.append(cells)

    if header:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{n + 1}"
    for i, w in enumerate(widths, start=1):
        # 中文字符实际占宽约为字母的两倍，粗略放大以免列被挤扁
        ws.column_dimensions[get_column_letter(i)].width = min(
            MAX_COL_WIDTH, max(MIN_COL_WIDTH, w + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_title(s: str) -> str:
    for ch in ':\\/?*[]':
        s = s.replace(ch, "_")
    return (s or "Sheet1")[:31]


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
